from flask import Flask, render_template, Response, request, redirect, url_for, send_from_directory
import cv2
from ultralytics import YOLO
import os
import time
from openpyxl import load_workbook, Workbook

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['STATIC_FOLDER'] = 'static'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['STATIC_FOLDER'], exist_ok=True)

model = YOLO("yolov8s-pose.pt")
video_atual = None
excel_file = "relatorio_velocidade.xlsx"

if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.append(["Data", "Hora", "ID", "Velocidade (m/s)", "Veredito"])
    wb.save(excel_file)

def obter_dados_excel():
    dados = []
    if os.path.exists(excel_file):
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            for row in list(ws.iter_rows(min_row=2, values_only=True))[-15:]:
                dados.append({'data': row[0], 'hora': row[1], 'id': row[2], 'vel': row[3], 'status': row[4]})
        except: pass
    return reversed(dados)

def gerar_frames(caminho_video):
    cap = cv2.VideoCapture(caminho_video)
    w, h = int(cap.get(3)), int(cap.get(4))
    fps = cap.get(cv2.CAP_PROP_FPS) or 30
    
    output_path = os.path.join('static', 'ultima_analise.mp4')
    fourcc = cv2.VideoWriter_fourcc(*'avc1') 
    out = cv2.VideoWriter(output_path, fourcc, fps, (w, h))

    linha_A_y, linha_B_y = int(h * 0.35), int(h * 0.85)
    dados_pessoas = {}

    while cap.isOpened():
        success, frame = cap.read()
        if not success: break
        
        frame_idx = cap.get(cv2.CAP_PROP_POS_FRAMES)
        results = model.track(frame, persist=True, verbose=False)
        
        # Desenha as linhas e o plot do YOLO ANTES de gravar no 'out'
        cv2.line(frame, (0, linha_A_y), (w, linha_A_y), (255, 255, 0), 2)
        cv2.line(frame, (0, linha_B_y), (w, linha_B_y), (0, 0, 255), 2)
        frame = results[0].plot() 

        if results[0].boxes.id is not None:
            ids = results[0].boxes.id.cpu().numpy().astype(int)
            boxes = results[0].boxes.xyxy.cpu().numpy().astype(int)
            for i, id_obj in enumerate(ids):
                y_pe = boxes[i][3]
                if id_obj not in dados_pessoas:
                    dados_pessoas[id_obj] = {'in': None, 'out': None}
                
                p = dados_pessoas[id_obj]
                if linha_A_y < y_pe < linha_B_y and p['in'] is None:
                    p['in'] = frame_idx
                elif y_pe > linha_B_y and p['in'] is not None and p['out'] is None:
                    p['out'] = frame_idx
                    t = (p['out'] - p['in']) / fps
                    vel = 10.0 / t if t > 0 else 0
                    status = "CORRENDO" if vel > 2.5 else "ANDANDO"
                    wb = load_workbook(excel_file); ws = wb.active
                    ws.append([time.strftime("%d/%m/%Y"), time.strftime("%H:%M:%S"), id_obj, round(vel, 2), status])
                    wb.save(excel_file)

        out.write(frame)
        ret, buffer = cv2.imencode('.jpg', frame)
        yield (b'--frame\r\n' b'Content-Type: image/jpeg\r\n\r\n' + buffer.tobytes() + b'\r\n')
               
    cap.release()
    out.release()

@app.route('/', methods=['GET', 'POST'])
def index():
    global video_atual
    if request.method == 'POST':
        file = request.files.get('video_file')
        if file:
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
            file.save(filepath); video_atual = filepath
            return redirect(url_for('index'))
    return render_template('index.html', dados=obter_dados_excel(), tem_video=(video_atual is not None))

@app.route('/video_feed')
def video_feed():
    return Response(gerar_frames(video_atual), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/download_video')
def download_video():
    return send_from_directory('static', "ultima_analise.mp4", as_attachment=True)

if __name__ == "__main__":
    app.run(debug=True)