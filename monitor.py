import cv2
import numpy as np
from ultralytics import YOLO
import time
import csv
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- CONFIGURAÇÕES ---
fonte = "videos/VID4.mp4" 
model = YOLO("yolov8s-pose.pt")
cap = cv2.VideoCapture(fonte)

w, h = int(cap.get(3)), int(cap.get(4))
fps = cap.get(cv2.CAP_PROP_FPS) or 30
# Criamos um vídeo mais largo para caber o Dashboard lateral
out = cv2.VideoWriter("analise.mp4", cv2.VideoWriter_fourcc(*'mp4v'), int(fps), (w + 400, h))

# --- PARÂMETROS ---
DISTANCIA_REAL_METROS = 10.0
linha_A_y, linha_B_y = int(h * 0.15), int(h * 0.95)
csv_file = "relatorio_velocidade.xlsx" # Vamos salvar direto em .xlsx para formatar
dados_pessoas = {}

# --- FUNÇÃO PARA FORMATAR O EXCEL ---
def formatar_excel(nome_arquivo):
    wb = load_workbook(nome_arquivo)
    ws = wb.active
    
    # Estilos
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    center_align = Alignment(horizontal="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in ws[1]: # Formatar cabeçalho
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2): # Formatar corpo da tabela
        for cell in row:
            cell.alignment = center_align
            cell.border = thin_border
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    wb.save(nome_arquivo)

# Criar arquivo inicial se não existir
if not os.path.exists(csv_file):
    wb = Workbook()
    ws = wb.active
    ws.append(["Data", "Hora", "ID", "Velocidade (m/s)", "Veredito"])
    wb.save(csv_file)

print(f"\n" + "="*50)
print(f"🚀 SISTEMA ANALÍTICO INICIADO")
print(f"📊 Relatório será salvo em: {csv_file}")
print("="*50 + "\n")

while cap.isOpened():
    ret, frame = cap.read()
    if not ret: break

    frame_atual = cap.get(cv2.CAP_PROP_POS_FRAMES)
    results = model.track(frame, persist=True, verbose=False)
    
    # Dashboard Lateral (Fundo Escuro)
    dash = np.zeros((h, 400, 3), dtype=np.uint8)
    dash[:] = (30, 30, 30) # Cinza escuro
    cv2.putText(dash, "LOG DE EVENTOS", (40, 60), cv2.FONT_HERSHEY_TRIPLEX, 1.2, (255, 255, 255), 2)
    cv2.line(dash, (40, 80), (360, 80), (100, 100, 100), 2)

    cv2.line(frame, (0, linha_A_y), (w, linha_A_y), (255, 255, 0), 2)
    cv2.line(frame, (0, linha_B_y), (w, linha_B_y), (0, 0, 255), 2)

    if results[0].boxes.id is not None:
        boxes = results[0].boxes.xyxy.cpu().numpy().astype(int)
        ids = results[0].boxes.id.cpu().numpy().astype(int)
        keypoints = results[0].keypoints.data.cpu().numpy()

        for i, id_obj in enumerate(ids):
            x1, y1, x2, y2 = boxes[i]
            cy = y2
            
            if id_obj not in dados_pessoas:
                dados_pessoas[id_obj] = {'in': None, 'out': None, 'c_run': 0, 'c_walk': 0, 'status': 'PROCESSANDO', 'vel': 0}
            
            p = dados_pessoas[id_obj]

            if linha_A_y < cy < linha_B_y:
                if p['in'] is None:
                    p['in'] = frame_atual
                    print(f"🟢 [ID {id_obj}] Entrou na zona de teste.")
                
                if len(keypoints) > i:
                    try:
                        diff = abs(keypoints[i][15][1] - keypoints[i][16][1])
                        if diff > (h * 0.04): p['c_run'] += 1
                        else: p['c_walk'] += 1
                    except: pass

            elif cy > linha_B_y and p['in'] is not None and p['out'] is None:
                p['out'] = frame_atual
                delta_t = (p['out'] - p['in']) / fps
                p['vel'] = DISTANCIA_REAL_METROS / delta_t
                p['status'] = "CORRENDO" if (p['c_run'] > p['c_walk'] or p['vel'] > 2.5) else "ANDANDO"
                
                # --- PRINT NO TERMINAL IGUAL ANTERIOR ---
                print(f"🏁 [ID {id_obj}] FINALIZOU:")
                print(f"   - Velocidade: {p['vel']:.2f} m/s")
                print(f"   - Veredito: {p['status']}")
                print("-" * 30)

                # --- SALVAR NO EXCEL ---
                agora = time.localtime()
                wb = load_workbook(csv_file)
                ws = wb.active
                ws.append([time.strftime("%d/%m/%Y", agora), time.strftime("%H:%M:%S", agora), 
                           id_obj, round(p['vel'], 2), p['status']])
                wb.save(csv_file)
                formatar_excel(csv_file)

            cor = (0, 0, 255) if p['status'] == "CORRENDO" else (0, 255, 0)
            cv2.rectangle(frame, (x1, y1), (x2, y2), cor, 2)
            cv2.putText(frame, f"ID {id_obj}", (x1, y1-10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, cor, 2)

    # Preencher Dashboard (Fonte aumentada e melhorada)
    y_pos = 150
    for id_p, info in list(dados_pessoas.items())[-10:]: # Mostra os últimos 10
        txt = f"ID {id_p}: {info['status']} ({info['vel']:.1f}m/s)"
        color = (100, 255, 100) if info['status'] == "ANDANDO" else (100, 100, 255)
        if info['vel'] == 0: color = (200, 200, 200)
        cv2.putText(dash, txt, (30, y_pos), cv2.FONT_HERSHEY_SIMPLEX, 0.8, color, 2)
        y_pos += 50

    # Combinar vídeo e dashboard
    final_frame = np.hstack((frame, dash))
    out.write(final_frame)
    cv2.imshow("Monitoramento de Corredor v3.0", final_frame)
    if cv2.waitKey(1) & 0xFF == ord('q'): break

cap.release()
out.release()
cv2.destroyAllWindows()
print(f"\n✅ Concluído! Relatório formatado salvo em: {csv_file}")